'''
Use: 
    robot.py simulation (s/r/sns/rns) exchange basic_curr-altcoin entry_price TP SL [limit_of_amount_to_sell] [sell_portion] 
    Example: 
    > python robot.py s btrx BTC-LTC 0.0017 0.0021 0.0015 100 
    Modes:
        s: simulation with stop-loss
        r: real mode with stop-loss
        sns: simulation and stop only on profit (no stop loss - not recommended)
        rns: real and stop only on profit (no stop loss - not recommended)
'''
################################ Libraries ############################################
# Standard libraries 
import os
import time
import sys
from sys import exit, argv
from time import localtime, strftime
import subprocess   
import math
import urllib2
import decimal
from decimal import Decimal
from openpyxl import Workbook, load_workbook   
from openpyxl.styles import Font, Fill
import json # requests
from shutil import copyfile # to copy files
import numpy as np
import traceback

# Decimal precision and roubding 
decimal.getcontext().prec = 25
decimal.getcontext().rounding = 'ROUND_DOWN'    

## Custom libraries 
from telegramlib import telegram                            # my lib to work with Telegram
from sqltools import query_lastrow_id, query        # proper requests to sqlite db
from loglib import logfile                                          # logging 
import platformlib as platform                                  # detecting the OS and setting proper folders 
import aux_functions               # various auxiliary functions   
aux_functions = aux_functions.aux_functions()

# Universal functions for all exchanges              
from exchange_func import (getticker, 
                                                  getopenorders, 
                                                  cancel, 
                                                  getorderhistory, 
                                                  getorder, 
                                                  getbalance, 
                                                  selllimit, 
                                                  getorderbook, 
                                                  buylimit, 
                                                  getbalances, 
                                                  binance_price_precise, 
                                                  binance_quantity_precise, 
                                                  getpositions, 
                                                  closepositions, 
                                                  bitmex_leverage, 
                                                  cancel_orders)

# Using coinigy to get prices so that there are no stringent restrictions on api request rates (frequency)
from coinigylib import coinigy 
coinigy = coinigy()

################################ Config - part I ############################################

### Import a configuration file 
import config 

### Price analysis library
import tdlib as tdlib


### Platform
platform = platform.platformlib()
platform_run, cmd_init, cmd_init_buy = platform.initialise() 
print "Initialising..."

### Set up the speedrun multiplier if need to test with higher speeds. Use any number - e.g. 1 is normal, 2 is 2x faster. 
speedrun = config.speedrun

### Telegram integration and preferred comm method 
chat = telegram()

### Command prompt parameters  

### Default values
no_input = False 
trailing_stop_flag = True  # default mode is to have trailing stop

### Input parameters 
try: 
    simulation_param = argv[1]
    if simulation_param == 's': 
        simulation = True
        stop_loss = True
    elif simulation_param == 'r': 
        simulation = False
        stop_loss = True
    elif simulation_param == 'sns': 
        simulation = True
        stop_loss = False
    elif simulation_param == 'rns': 
        simulation = False
        stop_loss = False
    elif simulation_param == 'rnts':
        trailing_stop_flag = False 
        simulation = False
        stop_loss = True
    else: 
        no_input = True 

    exchange_abbr = argv[2].lower()
    if exchange_abbr not in config.exch_supported: 
        print 'Incorrect exchange specified (should be btrx, bina, or bmex)\n\n'
        exit(0)
    
    # Commissions and full exchange names     
    if exchange_abbr == 'btrx': 
        exchange = 'bittrex' 
        comission_rate = config.comission_rate_bittrex
    elif exchange_abbr == 'bina': 
        exchange = 'binance' 
        comission_rate = config.comission_rate_binance
    elif exchange_abbr == 'bmex': 
        exchange = 'bitmex' 
        comission_rate = config.comission_rate_bitmex
    
    # Market to trade     
    market = argv[3].upper()
    try:
        trade, currency = market.split('-')
    except: 
        trade = market  # e.g. if only one market vs BTC is provided - such as XRPH18 on bitmex  
        currency = 'BTC'

    price_curr = float(argv[4])         # entry price 
    price_target = float(argv[5])      # target price 
    sl_target = float(argv[6])            # stop loss price
    price_entry = price_curr

    tp = round(price_target/price_curr, 5)
    sl = round(sl_target/price_curr, 5) 
    tp_p = (tp - 1.0)*100.0 
    sl_p = (1.0 - sl)*100.0 

    try:
        limit_sell_amount = float(argv[7])
    except: 
        limit_sell_amount = 0
    try:
        sell_portion = float(argv[8])
    except: 
        sell_portion = None    

except:
    no_input = True 

# Terminate if the input is improper 
if no_input:
    print '----------------------------------------------------------------------------------------------\n' + \
    'Run parameters not specified. Restart the script using:\n' + \
    'robot.py simulation (s/r/sns/rns) exchange basic_curr-altcoin entry_price TP SL [limit_of_amount_to_sell] [sell_portion]\n' +\
    'Example: > python robot.py s btrx BTC-LTC 0.0017 0.0021 0.0015 100\n\n' +\
    'Modes:\n>s (simulation with stop-loss)\n>r (real mode with stop-loss)\n>sns (simulation and stop only on profit)\n>rns (real and stop only on profit)'  
    exit(0) 
    
###  If simulation and parameters are not specified 
if simulation is True:
    if limit_sell_amount == 0: 
        limit_sell_amount = 100
    simulation_balance = limit_sell_amount
    sell_portion = limit_sell_amount

#### Gmail login and pass (if used) 
#fromaddr = config.fromaddr   
#toaddr = config.toaddr    
#email_passw = config.email_passw

################################ Config - part II ############################################
### Intervals and timers in seconds  

sleep_timer = config.sleep_timer                                      # Generic sleep timer. Applicable for the main monitoring loop and for the mooning procedure.
sleep_timer_buyback = config.sleep_timer_buyback      # Sleep timer for buybacks 
sleep_sale = config.sleep_sale                                          # Sleep timer for sell orders to be filled 
flash_crash_ind = config.flash_crash_ind                         # If something falls so much too fast - it is unusual and we should not sell (checking for 50% crashes)

## Interval and number of checks to get current (last) prices 
steps_ticker = config.steps_ticker  
sleep_ticker = config.sleep_ticker            

## Steps and timer for buybacks 
candle_steps = config.candle_steps         
candle_sleep = config.candle_sleep       

sleep_timer = int(sleep_timer/speedrun)
sleep_sale = int(sleep_sale/speedrun)
sleep_ticker = int(sleep_ticker/speedrun)
candle_steps = int(candle_steps/speedrun)

### To cancel buyback if there is an error and there were no sales made 
cancel_buyback = False 

### Bitmex margin 
bitmex_margin = config.bitmex_margin    # size of margin on bitmex, minor for now 

# Time analysis candles length 
# Possible options are in line with ohlc (e.g. 1h, 4h, 1d, 3d); customisable. See config for details.   
td_period = config.td_period   
td_period_extended = config.td_period_extended    
td_period_ext_opposite = config.td_period_ext_opposite    

# Market reference for BTC (see config) 
btc_market_reference = config.btc_market_reference

### Starting and default variables  
main_curr_from_sell = 0     
commission_total = 0        
alt_sold_total = 0  
decrease_attempts_total = 0  
value_original = 0
contracts_start = 0 
stopped_mode = '' 
short_flag = False # whether we are shorting, applicable for bitmex 
bitmex_sell_avg = 0 # for bitmex price averaging 
price_flip = False # for the confirmation of stops on the previous candle (should be a price flip there to stop, on td_period). False by default 
price_exit = None 
sl_extreme = None 
sale_trigger = False  
market_ref = None      # if we refer to a different exchange and market for td stats 
exchange_abbr_ref = None 
status_update = '' 

### Handle the reference to a different set of prices (from finex) in the case of usd-btc and bitmex 
if market == 'USD-BTC' and exchange == 'bitmex' and btc_market_reference:        # put in the config 
    market_ref = config.market_ref
    exchange_abbr_ref = config.exchange_abbr_ref
    print "Reference market {} on {}".format(market_ref, exchange_abbr_ref) 

### Strategy and thresholds for cases when there is no historical price data available
if currency in ['XMR', 'DASH', 'ETH', 'LTC', 'XMR']: 
    strategy = 'alt-med'
    diff_threshold = 0.045
elif currency == 'BTC': 
    strategy = 'btc'
    diff_threshold = 0.0255
else: 
    strategy = 'alt-volatile' 
    diff_threshold = 0.055

### Contingency for price-based analysis (see config) 
if strategy == 'btc': 
    var_contingency = config.var_contingency_btc
else: 
    var_contingency = config.var_contingency_alt

# Logger initialisation    
logger = logfile(market, 'trade')
 
    
##############################################  
##            Core get price / moon / sell functions           ##
##############################################
    
##################### Processing sell outcome results and generating messages 
def process_stat(status): 
    global exchange, market
    global db, cur, job_id
    global cancel_buyback 
    
    flag = True   # default flag returned
    
    if status == 'stop':
        message = 'Finishing up normally'
        flag = False
        sql_string = "UPDATE jobs SET selling = 0 WHERE job_id = {}".format(job_id)     # DB update
        rows = query(sql_string)

    if status == 'err_low': 
        message = 'Trade amount was too small and returned error, finishing up'
        aux_functions.send_notification(market, chat, 'Error: Too small trade', 'Too small trade to perform, finishing up')       
        cancel_orders(exchange, market)
        flag = False
        cancel_buyback = True 
        
    if status == 'no_idea': 
        message = 'Sell calls did not return proper answer, aborting'
        aux_functions.send_notification(market, chat, 'Error: No response from sell calls', 'Sell calls did not return proper answer, aborting')       
        cancel_orders(exchange, market)
        flag = False
        cancel_buyback = True 
        
    if status == 'abort_telegram': 
        message = 'Aborted as requested via Telegram'
        cancel_orders(exchange, market)
        flag = False
        cancel_buyback = True 
        
    return flag, message

##################### Checking if we need to stop buyback
def check_bb_flag():
    global market 
    global db, cur, bb_id
    
    sell_initiate = False 
    sql_string = "SELECT abort_flag FROM bback WHERE id = {}".format(bb_id)
    rows = query(sql_string)
    try: 
        bb_flag = rows[0][0] # first result 
    except: 
        bb_flag = 0  
    return bool(bb_flag)

##################### Looking for rebuy points (buyback), based on 4H candles price action or simpler price action depending on data availability
def buy_back(price_base): 
    global bb_id, market, exchange_abbr, exchange, sleep_timer_buyback
    global td_data_available, start_time, bars, strategy, time_bb_initiated # bars actually need to be recalculated as 1h is used for buyback
    global short_flag, td_period, td_period_extended, td_period_ext_opposite
    global market_ref, exchange_abbr_ref, diff_threshold, var_contingency
    global aux_functions, coinigy
    
    direction = None                        # to return information on the direction of the new detected trend 
    bars_check_avail = None 

    ### Greetings (for logs readability) 
    logger.lprint(["###################### BUY_BACK ###########################"])
    
    flag_reb_c = True 
    td_first_run = True 
    bback_result = False 

    if td_data_available != True: ## using a simple 5-min candles analysis if there is no 4H price data 
        price_l_arr = np.zeros(5)        #5x5-min candlesticks
        price_h_arr = np.zeros(5)       #5x5-min candlesticks
        crossed_arr = np.bool(5)        # for crossed
        logger.lprint([market, ': filling the price array'])
        
        bback_result_long = False 
        bback_result_short = False 
        
        # Filling the prices array with 5x5min candles
        while 0 in price_l_arr: 
            price_l, price_h, crossed_flag = coinigy.candle_analysis(exchange, exchange_abbr, market, logger, price_base)
            price_l_arr = np.append(price_l_arr, price_l)
            price_h_arr = np.append(price_l_arr, price_l)
            crossed_arr = np.append(crossed_arr, crossed_flag)
            price_l_arr = np.delete(price_l_arr, [0])
            price_h_arr = np.delete(price_h_arr, [0])
            crossed_arr = np.delete(crossed_arr, [0])
            # print "Lows", price_l_arr, '\nHighs', price_h_arr, '\nCrosses', crossed_arr   #DEBUG

        # Running until need to cancel 
        while flag_reb_c: 
            crossed_conf = (True in crossed_arr)                            # Any of candles should cross price_base   
            # LONGS check 
            lows_conf = aux_functions.equal_or_increasing(price_l_arr)                 # Higher or equal lows
            num_conf_long = ((price_h_arr >= price_base).sum()) >= 3     # At least 3 of 5 candles highs should be equal or above x
            bback_result_long = bool(lows_conf * crossed_conf * num_conf_long)
            logger.lprint([market, ": base", price_base, '| lows holding or higher', lows_conf, '| highs lower than base confirmation:', num_conf_long, '| crossed flag:', crossed_conf, '| result (long):', bback_result_long])
            # SHORTS check
            highs_conf = aux_functions.equal_or_decreasing(price_l_arr)                 # Lower or equal highs
            num_conf_short = ((price_l_arr <= price_base).sum()) >= 3     # At least 3 of 5 candles lows should be equal or below x
            bback_result_short = bool(highs_conf * crossed_conf * num_conf_short)
            logger.lprint([market, ": base", price_base, '| highs holding or lower', highs_conf, '| highs lower than base confirmation:', num_conf_short, '| crossed flag:', crossed_conf, '| result (short):', bback_result_short])
            
            # Check if we need to cancel 
            stop_bback = check_bb_flag()
            if stop_bback: 
                bback_result = False 
                flag_reb_c = False 
            
            # If we need to exit to proceed with buyback
            if bback_result_long: 
                bback_result = True
                direction = 'up' 
                logger.lprint([market, ": initiating buyback"])
                flag_reb_c = False 
            if bback_result_short: 
                bback_result = True
                direction = 'down' 
                logger.lprint([market, ": initiating buyback"])
                flag_reb_c = False 
                
            # Get new values 
            price_l, price_h, crossed_flag = coinigy.candle_analysis(exchange, exchange_abbr, market, logger, price_base)
            price_l_arr = np.append(price_l_arr, price_l)
            price_h_arr = np.append(price_l_arr, price_l)
            crossed_arr = np.append(crossed_arr, crossed_flag)
            price_l_arr = np.delete(price_l_arr, [0])
            price_h_arr = np.delete(price_h_arr, [0])
            crossed_arr = np.delete(crossed_arr, [0])
            
            # Updating DB
            if bb_id is not None: 
                sql_string = "UPDATE bback SET curr_price = {} WHERE id = {}".format(price_h, bb_id) 
                rows = query(sql_string)
               
            # Sleeping 
            time.sleep(sleep_timer_buyback)     
            
    ## If there is detailed 4H (or larger interval) data available (td_data_available) 
    else: 
        # Update to set stops according to 4H candles and TD 
        if td_first_run: 
            time_hour = time.strftime("%H")
            
        while flag_reb_c: 
            # Checking the need to update 
            time_hour_update = time.strftime("%H")
            if (time_hour_update <> time_hour) or td_first_run:
                # If this is the first run 
                if td_first_run: 
                    td_first_run = False 

                # Updating time 
                time_hour = time_hour_update
                # Updating TD values 
                td_info = tdlib.tdlib()
                bars = td_info.stats(market, exchange_abbr, td_period, 35000, 15, short_flag, market_ref, exchange_abbr_ref)     
                try: 
                    bars_extended = td_info.stats(market, exchange_abbr, td_period_extended, 60000, 15, short_flag, market_ref, exchange_abbr_ref)   
                    bars_check_avail = True 
                except: 
                    bars_check_avail = False 
                try: 
                    bars_ext_opposite = td_info.stats(market, exchange_abbr, td_period_ext_opposite, 80000, 15, short_flag, market_ref, exchange_abbr_ref)   
                    bars_check_avail = True 
                except: 
                    bars_check_avail = False 
                del td_info    
                
            # Check if we need to cancel 
            stop_bback = check_bb_flag()
            if stop_bback: 
                bback_result = False 
                flag_reb_c = False 
            
            # Checking time elapsed from the start of buyback 
            time_elapsed = (math.ceil(time.time() - time_bb_initiated ))/60    
            
            # Getting the current price and showing info on potential longs or potential shorts 
            price_upd = coinigy.get_avg_price(exchange, exchange_abbr, market, logger)
            
            if bars['td_direction'].iloc[-2] == 'up': #LONGS potential 
                logger.lprint([  exchange, market, "TD setup:", bars['td_setup'].iloc[-2], "| TD direction:", bars['td_direction'].iloc[-2], "4H candle high:", bars['high'].iloc[-2], "Current price:", price_upd, "Time elapsed (min):", time_elapsed  ])    
                if bars_check_avail and not config.ride_pullbacks: 
                    logger.lprint([  exchange, market, "TD setup (extended):", bars_extended['td_setup'].iloc[-2], "| TD direction:", bars_extended['td_direction'].iloc[-2] ])    
            elif (bars['td_direction'].iloc[-2] == 'down') and (exchange == 'bitmex'): #SHORTS potential, only for bitmex 
                logger.lprint([  exchange, market, "TD setup:", bars['td_setup'].iloc[-2], "| TD direction:", bars['td_direction'].iloc[-2], "4H candle low:", bars['low'].iloc[-2], "Current price:", price_upd, "Time elapsed (min):", time_elapsed  ])    
                if bars_check_avail and not config.ride_pullbacks: 
                    logger.lprint([  exchange, market, "TD setup (extended):", bars_extended['td_setup'].iloc[-2], "| TD direction:", bars_extended['td_direction'].iloc[-2] ])    
         
            # Updating DB
            if bb_id is not None: 
                sql_string = "UPDATE bback SET curr_price = {} WHERE id = {}".format(price_upd, bb_id) 
                rows = query(sql_string)
            
            ## CHECKING if we should reopen the position (buy back)  
            # Different strategy depending on the config.ride_pullbacks. 
            # A. If enabled - should happen quicker and no extended interval is checked 
            if config.ride_pullbacks: 
                # Check longs 
                if (bars['td_direction'].iloc[-2] == 'up') and (bars['td_direction'].iloc[-1] == 'up') and (time_elapsed > 10) and (price_upd > (bars['high'].iloc[-2])*(1 + var_contingency)):      # switching to long    
                    # This should _not_ be done on setup 7,8, or 9 
                    if (bars['td_setup'].iloc[-1] < 7):
                        bback_result = True 
                        direction = 'up'
                        flag_reb_c = False 
                        logger.lprint(["TD buyback initiated on the long side, riding the pullback"])
                  
                # Check shorts, only for bitmex
                if (bars['td_direction'].iloc[-2] == 'down') and (bars['td_direction'].iloc[-1] == 'down') and (time_elapsed > 10) and (price_upd < (bars['low'].iloc[-2])*(1 - var_contingency)) and (exchange == 'bitmex'):     
                    # This should _not_ be done on setup 7,8, or 9 
                    if (bars['td_setup'].iloc[-1] < 7):
                        bback_result = True 
                        direction = 'down'
                        flag_reb_c = False 
                        logger.lprint(["TD buyback initiated on the short side, riding the pullback"])
                                
            # B. If we do not want to ride pullbacks 
            else: # Checking the larger interval and more stringent conditions           
                # Check longs 
                if (bars['td_direction'].iloc[-2] == 'up') and (bars['td_direction'].iloc[-1] == 'up') and (time_elapsed > 30) and (price_upd > (bars['high'].iloc[-2])*(1 + var_contingency)):      # switching to long    
                    # This should _not_ be done on setup 7,8, or 9 
                    if (bars['td_setup'].iloc[-1] < 7):
                        # Depending on the short flag, selecting the interval  
                        if not short_flag:  # same direction - checking bars_extended 
                            bars_check = bars_extended
                        else:   # different direction - checking a larger period 
                            bars_check = bars_ext_opposite
                        # Checking the conditions                 
                        if (bars_check_avail and bars_check['td_direction'].iloc[-1] == 'up') or (bars_check_avail == False): 
                            bback_result = True 
                            direction = 'up'
                            flag_reb_c = False 
                            logger.lprint(["TD buyback initiated on the long side"])
                            if bars_check_avail == False: 
                                logger.lprint(["Note that higher - timeframe TD analysis is not available"])    
                
                # Check shorts, only for bitmex
                if (bars['td_direction'].iloc[-2] == 'down') and (bars['td_direction'].iloc[-1] == 'down') and (time_elapsed > 30) and (price_upd < (bars['low'].iloc[-2])*(1 - var_contingency)) and (exchange == 'bitmex'):     
                    # This should _not_ be done on setup 7,8, or 9 
                    if (bars['td_setup'].iloc[-1] < 7):
                        # Depending on the short flag 
                        if short_flag:  # same direction - checking bars_extended 
                            bars_check = bars_extended
                        else:     # different direction - checking a larger period 
                            bars_check = bars_ext_opposite
                        # Checking the conditions      
                        if (bars_check_avail and bars_check['td_direction'].iloc[-1] == 'down') or (bars_check_avail == False): 
                            bback_result = True 
                            direction = 'down'
                            flag_reb_c = False 
                            logger.lprint(["TD buyback initiated on the short side"])
                            if bars_check_avail == False: 
                                logger.lprint(["Note that higher - timeframe TD analysis is not available"])    
            
            # Sleeping 
            time.sleep(sleep_timer_buyback)     
            
    # Finishing up 
    return bback_result, direction
    

##################### Update information on performed orders
def sell_orders_info():
    global simulation, main_curr_from_sell, commission_total, alt_sold_total, orders_start, no_sell_orders, market, limit_sell_amount
    global exchange, exchange_abbr, logger 
    global price_exit, bitmex_sell_avg
    
    # Updating order history to collect information on new orders (disable in the simulation mode)
    # Further speed improvement would be to change the structure to a proper dict here right away    
    
    try: # to handle errors 
        if simulation != True: 
            # Reset values if we are not simulating
            main_curr_from_sell = 0     
            commission_total = 0        
            alt_sold_total = 0 
            
            # Getting information on _sell_ orders executed
            orders_opening_upd = getorderhistory(exchange, market) 
            for elem in orders_opening_upd: 
                orders_new.add(elem['OrderUuid'])
            orders_executed = orders_new.symmetric_difference(orders_start) 
     
            if orders_executed == set([]):
                logger.lprint(["No sell orders executed"])
                no_sell_orders = True 
            else:
                logger.lprint(["New executed orders"])  
                
                for elem in orders_executed: 
                    order_info = getorder(exchange, market, elem)               
                    if exchange != 'bitmex': 
                        main_curr_from_sell += order_info['Price']  
                        commission_total += order_info['CommissionPaid']
                        qty_sold = order_info['Quantity'] - order_info['QuantityRemaining'] 
                        alt_sold_total += qty_sold                            
                        logger.lprint([">", elem, "price", order_info['Price'], "quantity sold", qty_sold ]) #DEBUG 
                    else: 
                        price_exit = bitmex_sell_avg
                        if price_exit != 0: 
                            if market == 'USD-BTC': 
                                main_curr_from_sell = contracts_start/price_exit
                            else: 
                                main_curr_from_sell = contracts_start*price_exit
                logger.lprint(["Total price", main_curr_from_sell, "alts sold total", alt_sold_total]) #DEBUG
        else:
            # If the simulation is True - main_curr_from_sell will have simulated value and the commission would be zero. Updating quantity. 
            alt_sold_total = limit_sell_amount
            price_exit = coinigy.get_avg_price(exchange, exchange_abbr, market, logger)
            if exchange == 'bitmex': 
                if market == 'USD-BTC': 
                    main_curr_from_sell = contracts_start/price_exit
                else: 
                    main_curr_from_sell = contracts_start*price_exit

    except: 
        err_msg = traceback.format_exc()
        comm_string = 'Could not het sell orders history from {} on {}. Reason: {}. Check the results'.format(market, exchange, err_msg)
        logger.lprint([comm_string])    
        chat.send(comm_string)  
    
##################### Sell orders outcome 
def sell_orders_outcome():
    global no_sell_orders, total_gained, main_curr_from_sell, value_original, commission_total, total_gained_perc, market
    global price_exit, contracts_start # to use in buyback
    global short_flag
    
    emoji_text = ''
    
    if no_sell_orders != True: 
        # Calculating totals 
        total_gained = float(main_curr_from_sell) - float(value_original) - float(commission_total)
        
        # Here division by zero error handling
        if float(value_original)  != 0: 
            total_gained_perc = 100*float(total_gained)/float(value_original)   
        else: 
            total_gained_perc = 0 
        
        # Depending on the trade direction 
        if (short_flag and (total_gained_perc < 0)) or (not short_flag and (total_gained_perc >= 0)): 
            txt_result = 'gained'
        else: 
            txt_result = 'lost'  
        
        # Average exit price (value/quantity)
        if exchange != 'bitmex':   
            price_exit = float(main_curr_from_sell)/float(alt_sold_total)
        
        ''' 
        # Bitmex: Does not work due to data limitations - revise using tradehistory
        # price_exit = float(contracts_start)/float(main_curr_from_sell)   # for bitmex, calculation is done through contracts    # commented until further fix
        ''' 
        #print  "Price exit calc: price_exit {}, contracts_start {}, main_curr_from_sell {}".format(price_exit, contracts_start, main_curr_from_sell) # DEBUG 
            
        percent_gained = str(round(total_gained_perc, 2))
        trade_time = strftime("%Y-%m-%d %H:%M", localtime())
        
        logger.lprint(['Total from all sales', main_curr_from_sell, 'total commission', commission_total])
        logger.lprint(['Profit ', total_gained, ':', round(total_gained_perc, 2), '%']) 
        
        # Send the notification about results    
        # To improve: if bitmex is used, margin should be accounted for 
        
        # Emoji to use 
        if (short_flag and (price_exit < price_entry)) or (not short_flag and (price_exit > price_entry)):
            emoji_text = '\xF0\x9F\x9A\x80'         # rocket    
        if (short_flag and (price_exit > price_entry)) or (not short_flag and (price_exit < price_entry)):
            emoji_text = '\xF0\x9F\x90\xA3'         # chicken    
        
        ''' # Fix the calculations 
        msg_result = '{} {}: Total {} gained from all sales: {}. Commission paid: {}. Trade outcome: {} % {}. \nEntry price: {}, exit price: {}, short_flag {}'.format(emoji_text, market, str(trade), main_curr_from_sell, str(commission_total),  str(percent_gained), txt_result, str(price_entry), str(price_exit), short_flag)        
        ''' 
        msg_result = '{} {}: Entry price: {}, exit price: {}, short_flag {}'.format(emoji_text, market, str(price_entry), str(price_exit), short_flag)        
        aux_functions.send_notification(market, chat, 'Finished', msg_result) 
        
        # Update the xls register 
        try:
            wb = load_workbook(config.trade_hist_filename)
            ws = wb['BOT']
            new_line = [trade_time, trade, currency, alt_sold_total, price_curr, price_exit, main_curr_from_sell, total_gained, percent_gained, simulation]
            ws.append(new_line)
            max_row = ws.max_row
            # Apply a style 
            index_row = "{}:{}".format(max_row, max_row) 
            for cell in ws[index_row]:
                cell.font = Font(name='Arial', size=10)
            wb.save(config.trade_hist_filename)
            
            #if platform_run != 'Windows':  #uncomment if needed 
            #    copyfile('/home/illi4/Robot/Trade_history.xlsx', '/mnt/hgfs/Shared_folder/Trade_history.xlsx')
            
        except: 
            logger.lprint(['Trade history xls unavailable']) 

##################### Setting stop loss based on price data
def stop_reconfigure(mode = None): 
    global db, cur, job_id
    global time_hour, time_hour_comms
    global market, exchange_abbr, strategy 
    global price_entry, short_flag, td_period
    global var_contingency
    global bars_4h 
    global market_ref, exchange_abbr_ref
    global status_update
    global rsi_1h, rsi_4h 
    
    price_flip_upd = None # default is none so that we do not update it every time the stop_reconfigure is called 
    price_direction_move = None 
    sl_target_upd = None 
    sl_upd = None 
    sl_p_upd = None  
    sl_extreme_upd = None # for the absolute min / max of TD setup 
    
    td_info = tdlib.tdlib()
    
    time_hour_update = time.strftime("%H")
 
    if (time_hour_update <> time_hour) or mode == 'now': 
        
        # Updating the current hour and the TD values 
        time_hour = time_hour_update
        bars_4h = td_info.stats(market, exchange_abbr, td_period, 35000, 15, short_flag, market_ref, exchange_abbr_ref)     

        # New logic: return the TD direction of the last candle per td_interval 
        price_direction_move = bars_4h['td_direction'].iloc[-1]                         # returns 'up' or 'down' 
        price_direction_move_previous = bars_4h['td_direction'].iloc[-2]        # returns 'up' or 'down' 
        #print "CHECK: short flag", short_flag, "price_direction", price_direction_move   #DEBUG
        
        # Update the RSI values 
        rsi_4h = td_info.stats_rsi_only(market, exchange_abbr, '4h', 35000, 15, short_flag, market_ref, exchange_abbr_ref)[-1]     
        rsi_1h = td_info.stats_rsi_only(market, exchange_abbr, '1h', 35000, 15, short_flag, market_ref, exchange_abbr_ref)[-1]      
        
        # We will be considering that there is a price flip if we have a candle in setup with different colour which is followed by the same colour 
        # So the rule will be for example if we are long and there is bearish flip, then there are 1 and 2 red -> price_flip_upd is true
        if ((not short_flag and price_direction_move == 'down' and price_direction_move_previous == 'down') 
        or (short_flag and price_direction_move == 'up' and price_direction_move_previous == 'up')): 
            price_flip_upd = True 
            #print "> Price_flip_upd", price_flip_upd #DEBUG 

        if not short_flag: # the position is long 
            sl_target_upd = bars_4h['low'].iloc[-2] * (1 - var_contingency)   
            sl_upd = round(sl_target_upd/price_entry , 5) 
            sl_p_upd = (1.0 - sl_upd)*100.0 
            if bars_4h['move_extreme'].iloc[-1]  is not None: 
                sl_extreme_upd = bars_4h['move_extreme'].iloc[-1] * (1 - var_contingency)      
                
        else: # the position is short 
            sl_target_upd = bars_4h['high'].iloc[-2] * (1 + var_contingency)   
            sl_upd = round(sl_target_upd/price_entry , 5) 
            sl_p_upd = (1.0 + sl_upd)*100.0  
            if bars_4h['move_extreme'].iloc[-1]  is not None: 
                sl_extreme_upd = bars_4h['move_extreme'].iloc[-1] * (1 + var_contingency)     
        
        logger.lprint([  "New stop loss level based on the last candle: {}, setup direction: {}. Flip: {}".format(sl_target_upd, price_direction_move, price_flip_upd) ])
        logger.lprint([  "New extreme stop value:", sl_extreme_upd ])
        logger.lprint([  "Price flip analysis: {}, direction current: {}, direction previous: {}, short flag: {}".format(price_flip_upd, price_direction_move,
        price_direction_move_previous, short_flag) ])
        #print "> Returning price_flip_upd {}, sl_target_upd {}, sl_upd {}, sl_p_upd {}, sl_extreme_upd {}".format(price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd)  #DEBUG
    
    # Status updates 
    if ((int(time_hour_update) - int(time_hour_comms)) == 4) or mode == 'now': 
        time_hour_comms = time_hour_update 
        status_update = "Status update | {} {}: \nextreme_move stop {} \n4h-based stop {} \nprice flip confirmation {}".format(market, exchange_abbr, sl_extreme_upd, sl_target_upd, price_flip_upd)
        status_update += "\nTD: \ncurrent {} {} \nprevious {} {}".format(bars_4h['td_setup'].iloc[-1], bars_4h['td_direction'].iloc[-1], bars_4h['td_setup'].iloc[-2], bars_4h['td_direction'].iloc[-2])
        status_update += "\nRSI: 4H {:.2f}, 1H {:.2f}".format(rsi_4h, rsi_1h)
        chat.send(status_update)
    
    # Updating the db with the current SL value 
    if sl_target_upd is not None: 
        sql_string = "UPDATE jobs SET sl={}, sl_p={} WHERE job_id={}".format(sl_target_upd, sl_upd, job_id)
        rows = query(sql_string)   
    
    del td_info
    
    return price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd
            
##################### Mooning trajectory procedure
### Currently works in the same way as just the main cycle when TD price data is available (simply reconfiguring stops dynamically) 
### Thus, there is no need to use it on price-based analysis
       
def to_the_moon(price_reached):     
    # Global variables used 
    global main_curr_from_sell, value_original, price_curr, commission_total, price_target, t_m_id, approved_flag, offset_check, comission_rate
    global sleep_timer
    global db, cur, job_id
    global stopped_price
    global trailing_stop_flag, start_time, bars, strategy, diff_threshold
    global sl, sl_target, sl_p
    global short_flag, price_flip, sl_extreme
    global bars_4h, var_contingency
    global exchange, exchange_abbr, market, logger

    sale_trigger = False # default
    
    # Thresholds for post-profit fallback for BTC or ALTS, when detailed price data is not available 
    post_sl_level = (1 - var_contingency)    
    price_max = price_reached                                       # this will be changed mooning forward
    price_cutoff = price_reached * post_sl_level          # to sell on original TP if we fall below 
    if td_data_available: 
        trailing_stop = sl_target 
    else: 
        trailing_stop = price_max * post_sl_level            # to sell on new high * stop loss threshold   
    
    logger.lprint(["Mooning from:", price_max])   
    rocket_flag = True
    
    # Running the loop 
    while rocket_flag:  
        # Update to set stops according to 4H candles and TD 
        if td_data_available: 
            price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd = stop_reconfigure()
            #print ">>> Returned price_flip {}, sl_target_upd {}, sl_upd {}, sl_p_upd {}, sl_extreme_upd {}".format(price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd)  #DEBUG
            if sl_target_upd is not None: 
                trailing_stop = sl_target_upd
                sl = sl_upd
                sl_p = sl_p_upd    
            if sl_extreme_upd is not None: 
                sl_extreme = sl_extreme_upd
            if price_flip_upd is not None: 
                price_flip = price_flip_upd
            
        price_last_moon = coinigy.get_avg_price(exchange, exchange_abbr, market, logger)
        increase_info = 100*float(price_last_moon - price_target)/float(price_target) 
        logger.lprint(["Price update:", price_last_moon, "in comparison with the original target:", round(increase_info, 2), "%"])

        # Updating the db 
        sql_string = "UPDATE jobs SET price_curr={}, percent_of={}, mooning={} WHERE job_id={}".format(round(price_last_moon, 8), str(round(increase_info, 2)), 1, job_id)
        rows = query(sql_string)
        
        # Depending on whether there is short or long 
        if ((short_flag != True) and (price_last_moon > price_max)) or ((short_flag == True) and (price_last_moon < price_max)):  
            # Setting higher thresholds if there is no 4H data
            price_max = price_last_moon
            if not td_data_available: 
                trailing_stop = price_max * post_sl_level        
            logger.lprint(["Last price:", price_max, "| trailing stop", trailing_stop, "| original take profit", price_cutoff])

        #  Checking if this is a time to sell now   
        #  starting only when trailing_stop_flag is active (should not be doing this for BTC runs) 
        # print ">> Price last moon (to compare)", price_last_moon, "maximum price", price_max, "price_cutoff", price_cutoff, "trailing_stop", trailing_stop  # DEBUG # 
        
        if trailing_stop_flag: 
            # Simplified this back to basics as we are using the 4H rule and selling if we are falling behind the bullish candle 
            # Depending on the long or short 
            if (((not short_flag) and price_flip and (price_last_moon <= min(price_cutoff, trailing_stop)) )  # if we are long and the price drops below original or trailing stop 
            or (short_flag and price_flip and (price_last_moon >= max(price_cutoff, trailing_stop))) ):  
                logger.lprint(["Run out of fuel @", price_last_moon])
                # Check if we need to sell. No need to do this if we have price action data (backtested for performance) 
                if not td_data_available: 
                    sale_trigger = ensure_sale(price_last_moon)   
                else: 
                    sale_trigger = True 
                logger.lprint(["Sale trigger (post-profit)", sale_trigger])
            
            # Now checking sale trigger and selling if required         
            if sale_trigger == True:  
                logger.lprint(["Triggering trailing stop on", price_last])
                aux_functions.send_notification(market, chat, 'Sell: Post-TP', exchange + ' : ' + market + ': Triggering trailing stop on the level of ' + str(price_last))
                status = sell_now(price_last_moon)
                # Update the status
                rocket_flag, stat_msg = process_stat(status)
                logger.lprint([stat_msg])            
                # For buyback - using rebuy price
                if short_flag:  
                    stopped_price = min(price_cutoff, trailing_stop)
                else: 
                    stopped_price = max(price_cutoff, trailing_stop)
                            
        # Check if 'sell now' request has been initiated
        sell_init_flag = check_sell_flag()
        if sell_init_flag == True:       
            logger.lprint(["Sale initiated via Telegram @", price_last])
            status = sell_now(price_last_moon)
            sql_string = "UPDATE jobs SET selling = 0 WHERE job_id = {}".format(job_id)     # updating the DB 
            rows = query(sql_string)
            
            # Handling results
            rocket_flag, stat_msg = process_stat(status)
            logger.lprint([stat_msg])
            # For buyback - using rebuy price
            if short_flag:  
                stopped_price = min(price_cutoff, trailing_stop)
            else: 
                stopped_price = max(price_cutoff, trailing_stop)
                
        # Checking Telegram requests and answering 
        if rocket_flag:
            approved_flag = check_cancel_flag()
            if approved_flag == False: 
                logger.lprint(["Shutdown was requested via Telegram"])   
                sleep_timer = 0
            time.sleep(sleep_timer)

        if approved_flag == False:  # aborting if asked          
            status = 'abort_telegram'
            rocket_flag, stat_msg = process_stat('abort_telegram')

    # Finished the loop - returning the proper code
    return status

##################### Anti-manipulation and anti-flash-crash filter for cases when we do not rely on price time analysis  
def ensure_sale(check_price): 
    global short_flag
    global exchange, exchange_abbr, market, logger, price_base
    
    proceed_sale = False           
    price_arr = np.zeros(3)          # 3 * N-min candlesticks  (see candle_extreme for N) 
    logger.lprint(["Running ensure_sale check"])
    
    ## Filling the prices array - will be checking for lower highs 
    while (0 in price_arr):  
        approved_flag = check_cancel_flag()  # checking Telegram requests and answering 
        if approved_flag == False: 
            break
    
        price_lowest, price_highest, crossed_flag_info = coinigy.candle_analysis(exchange, exchange_abbr, market, logger, check_price) 
        #candle_extreme('H')  
        
        if short_flag != True: # LONGS  
            price_arr = np.append(price_arr, price_highest)
        else:  # SHORTS   
            price_arr = np.append(price_arr, price_lowest)
        price_arr = np.delete(price_arr, [0])
        
        # Selling on the series of lower or same highs of 3 x N-min candlesticks when the price array is filled for longs: 
        if short_flag != True: # LONGS  
            if (0 not in price_arr): 
                logger.lprint(["High in the candle:", price_highest, "| lower or same highs:", aux_functions.equal_or_decreasing(price_arr)])  #logger.lprint([price_arr]) # DEBUG
                if aux_functions.equal_or_decreasing(price_arr): 
                    proceed_sale = True
                    break
            else: 
                logger.lprint(["High in the candle:", price_highest])  #logger.lprint([price_arr]) # DEBUG
                
            # If we are back above the check_price value - exit the cycle and return false 
            if price_highest > check_price: 
                logger.lprint(["Cancelling ensure_sale since the price is back to normal"])  
                proceed_sale = False
                break
                
        else: # SHORTS      
            if (0 not in price_arr): 
                logger.lprint(["Low in the candle:", price_lowest, "| higher or same lows:", aux_functions.equal_or_increasing(price_arr)])  #logger.lprint([price_arr]) # DEBUG
                if aux_functions.equal_or_increasing(price_arr): 
                    proceed_sale = True
                    break
            else: 
                logger.lprint(["Low in the candle:", price_lowest])  #logger.lprint([price_arr]) # DEBUG
                
            # If we are back above the check_price value - exit the cycle and return false 
            if price_lowest < check_price: 
                logger.lprint(["Cancelling ensure_sale since the price is back to normal"])  
                proceed_sale = False
                break

    return proceed_sale     

##################### Anti-manipulation and anti-flash-crash filter for TD price action available 
def ensure_td_sale(check_price): 
    global short_flag
    global exchange, exchange_abbr, market, logger, market_ref
    
    # Checking the last 2 closes of 10-min candles (excluding the current one) 
    proceed_sale = False       
    logger.lprint(["Running ensure_sale check for TD price"])
    
    td_info = tdlib.tdlib()
    
    bars_10min = td_info.stats(market, exchange_abbr, '10min', 1000, 5, short_flag, market_ref, exchange_abbr_ref)     
    logger.lprint([ "Last two 10-min candles close values: {}, {}".format(bars_10min['close'].iloc[-3], bars_10min['close'].iloc[-2]) ])
    
    # Selling only if the last 2 closes went beyond our threshold 
    if not short_flag and (bars_10min['close'].iloc[-2] < check_price) and (bars_10min['close'].iloc[-3] < check_price): 
        proceed_sale = True 
    if  short_flag and (bars_10min['close'].iloc[-2] > check_price) and (bars_10min['close'].iloc[-3] > check_price): 
        proceed_sale = True 
    
    # Free up memory 
    del bars_10min    
    del td_info
    
    return proceed_sale     
    
##################### Main sell function to sell at current prices   
# Will be performed until the balance available for sale is zero or slightly more      
def sell_now(at_price):

    global bitmex_sell_avg  # for bitmex average price calc 
    bitmex_sell_avg_arr = []
    
    # To decrease price gradually compared to the last average sell price if orders are not filled. Start with zero (percent), maximum 5%
    decrease_price_step = 0.0 
    decrease_attempts_total = 0 
    # First run flag now to sleep on the first call 
    proceed_w_sleep = False
    
    # Global variables used 
    global main_curr_from_sell, value_original, price_curr, commission_total, simulation, currency, market, t_m_id
    global approved_flag, offset_check, simulation_balance, sell_portion, limit_sell_amount, comission_rate, exchange
    global sleep_sale, steps_ticker, sleep_ticker, balance_start, contracts_start, short_flag
    global db, cur, job_id
    global chat
    global exchange, exchange_abbr, logger
    
    # Starting balance for further use. Should be done with all orders cancelled
    cancel_orders(exchange, market)
    
    # Get balance
    if simulation != True: 
        balance = getbalance(exchange, currency)
        balance_start  = Decimal('{0:.8f}'.format(balance['Available']))   # to correctly work with decimal numbers; not needed for bitmex 

        if exchange != 'bitmex':         
            logger.lprint(["Balance available to sell", balance_start])    
    
    if limit_sell_amount is not None: 
        limit_sell_amount = Decimal(str(limit_sell_amount))     # using str, we will not have more decimal numbers than needed
    if sell_portion is not None: 
        sell_portion = Decimal(str(sell_portion))  
    
    if simulation == True: 
        balance_start = Decimal(str(simulation_balance))
        balance_available = Decimal(str(simulation_balance))
        remaining_sell_balance = Decimal(str(simulation_balance))
        
    # Limiting if required. Should be done with orders cancelled
    if (limit_sell_amount < balance_start) and (limit_sell_amount > 0):
        balance_adjust = Decimal(str(balance_start)) - Decimal(str(limit_sell_amount))
        balance_start = Decimal(str(limit_sell_amount))
        #print ">> Adjust", balance_adjust, "Bal_start", balance_start, "Limit sell am", limit_sell_amount      #DEBUG 
        logger.lprint(["Limiting total amount to be sold. Total:", limit_sell_amount, "Adjustment:", balance_adjust])
    else:
        balance_adjust = 0

    # For bitmex, we will be trading contracts, no adjustments are available. Getting the balances and setting the original value 
    if exchange == 'bitmex': 
        if simulation != True: 
            # There were issues with testnet returning blanks so changed this 
            contracts_check = {}
            positions = getpositions(exchange, market)  # first not empty result 
            for position in positions: 
                if position != {}: 
                    contracts_check = position 
                    break # exit the for loop 
            print 'contracts_check', contracts_check #TEST 
            # If nothing was found  
            if contracts_check == {}: 
                sell_run_flag = False
                contracts = 0
            else: 
                if market == 'USD-BTC': 
                    contracts = contracts_check['contracts'] 
                    value_original = Decimal(str(contracts_check['contracts_no']))
                else: 
                    contracts = contracts_check['contracts_no'] 
                    value_original = Decimal(str(contracts))*Decimal(price_entry)  
               
                contracts_start = contracts
                balance_available = contracts
                balance_adjust = 0 
                sell_portion = balance_available
             
        else: # if we are in the simulation mode 
            contracts =  price_entry * simulation_balance    #get_last_price(market) * simulation_balance
            contracts_start = contracts
            value_original = simulation_balance
    else: # for other exchanges     
        value_original = Decimal(str(price_entry)) * balance_start    
 
    logger.lprint(["Original value:", value_original])
    
    # Main sell loop
    sell_run_flag = True
    stopmessage = 'stop' # default stop message meaning successful sale
    
    while sell_run_flag: 
        decrease_price_flag = False     # Reset the overall flag to decrease price 
      
        # Wait until existing orders are cancelled - that is why we need sleep here and not in the end 
        # Checking Telegram requests and cancelling if needed
        if proceed_w_sleep: 
            time.sleep(sleep_sale)
        
        # 0. Check open orders, cancel if unfilled, and decrease price further compared to average last 
        my_orders = getopenorders(exchange, market)
        if my_orders <> '': 
            for val in my_orders:
                # Checking if some are open not filling
                if (val['Quantity'] == 0):
                    unfilled_prop = 0
                else:
                    unfilled_prop = Decimal(str(val['QuantityRemaining']))/Decimal(str(val['Quantity']))
                if unfilled_prop >= 0.05:  # if more than 5% still left in the order
                    logger.lprint(["Cancelling unfilled order:", val['OrderUuid'], "quantity", val['Quantity'], 
                           "quantity remaining", val['QuantityRemaining'], "limit", val['Limit'], "price", val['Price']
                           ]) 
                    cancel_stat = cancel(exchange, market, val['OrderUuid'])
                    time.sleep(5) # Wait for cancellations to be processed just in case 
                    # Then we will get information on available balance which includes cancellations
                    # Set decrease price flag
                    decrease_price_flag = True
                    decrease_attempts_total += 1

        # Decrease price more compared to last prices if required
        if (decrease_price_step < 0.0255) and decrease_price_flag:     
            if short_flag != True: #LONG
                decrease_price_step += 0.001
                logger.lprint(["Sell price will be decreased on", decrease_price_step*100, "%"]) 
            else: #SHORT 
                decrease_price_step -= 0.001
                logger.lprint(["Sell price will be increased on", decrease_price_step*100, "%"]) 
            
        # Notify if a position cannot be sold for a long time 
        if decrease_attempts_total >= 30: 
            time_passed = int(decrease_attempts_total*(sleep_sale + steps_ticker*sleep_ticker)/60)
            logger.lprint(["Unable to sell the position for more than", time_passed, "minutes"]) 
            chat.send(market +": unable to sell the position for more than " + time_passed + " minutes")
                        
        # 1. Get the available balance and proceed with selling       
        if simulation != True: 
            balance = getbalance(exchange, currency)
            balance_available = Decimal('{0:.8f}'.format(balance['Available']))
            # print ">> Balance_available", balance_available #DEBUG 
        else:
            # If we are in the simulation mode - use the value from the previous run
            balance_available = remaining_sell_balance           
        
        # For bitmex, we will be trading contracts, no adjustments are available 
        if exchange == 'bitmex': 
            # There were issues with testnet returning blanks so changed this 
            contracts_check = {}
            positions = getpositions(exchange, market)  # first not empty result 
            for position in positions: 
                if position != {}: 
                    contracts_check = position 
                    break # exit the for loop 
            # If nothing was found 
            if contracts_check == {}: 
                sell_run_flag = False
            else: 
                if market == 'USD-BTC': 
                    contracts = contracts_check['contracts'] 
                else: 
                    contracts = contracts_check['contracts_no'] 
                balance_available = contracts
                balance_adjust = 0 
                sell_portion = balance_available
        else: # for the other exchanges 
            #print ">> Balance_available pre", balance_available    #DEBUG  
            #print  ">> Balance_adjust pre", balance_adjust     #DEBUG  
            
            # Adjusting according to the limit 
            balance_available = balance_available - Decimal(str(balance_adjust))
            if sell_portion == None: 
                sell_portion = balance_available           

        # Check if we have sold everything 
        if balance_available <= balance_start * Decimal(0.01):
            sell_run_flag = False
        
        # Error strings for exchanges 
        err_1 = 'DUST_TRADE_DISALLOWED_MIN_VALUE_50K_SAT'
        err_2 = 'MIN_TRADE_REQUIREMENT_NOT_MET'
        
        # 2. If something is still required to be sold
        if sell_run_flag: 
            logger.lprint(["Order amount", balance_available, "at price threshold", at_price, "split on", sell_portion])
            remaining_sell_balance = balance_available   
            if exchange == 'bitmex': 
                sale_steps_no = 1       # for the whole position (at least for now) 
            else: 
                sale_steps_no = int(math.ceil(round(Decimal(str(balance_available))/Decimal(str(sell_portion)), 3)))   
            #print ">> Sell amount", balance_available, "remaining_sell_balance", remaining_sell_balance  #DEBUG#
            
            # Selling loop 
            for i in range(1, sale_steps_no + 1):                
                # Check how much should we sell at this step
                if sell_portion > remaining_sell_balance: 
                    sell_q_step = remaining_sell_balance
                else:
                    sell_q_step = sell_portion
                
                # Price update
                if exchange != 'bitmex': 
                    price_last_sell = coinigy.get_avg_price(exchange, exchange_abbr, market, logger)
                else: 
                    # When we are long, on the exit we sell -> get the price from bids (the highest which is the first in the array)
                    # When we are short, on the exit we buy -> get the price from asks (the lowest, which is the first in the array)
                    if not short_flag: #LONG
                        price_last_sell = float(getorderbook('bitmex', market, 'bids')[0]['Rate'])
                    else: # SHORT   
                        price_last_sell = float(getorderbook('bitmex', market, 'asks')[0]['Rate'])
                
                # Decreasing the price if necessary
                price_to_sell = price_last_sell*(1 - decrease_price_step)
                logger.lprint(["Placing SELL order: Q:", sell_q_step, "@", price_to_sell, "Last market price:", price_last_sell, 
                       "Remaining balance after sale:", round(remaining_sell_balance - sell_q_step, 6)])
                
                # Actually place sell orders if we are not in the simulation mode - re-check
                if simulation != True: 
                    # For bitmex, we will be placing contracts in the other direction (short)
                    if exchange == 'bitmex': 
                        # Balance_available is the number of contracts here. Creating orders depending on the side (long or short) 
                        if market == 'USD-BTC': 
                            price_to_sell = round(price_to_sell, 0)
                        else: 
                            price_to_sell = round(price_to_sell, 20)
                        bitmex_sell_avg_arr.append(price_to_sell) 
                        
                        if short_flag != True: #LONG
                            sell_result = selllimit(exchange, market, sell_q_step, price_to_sell, balance_available) 
                        else: # SHORT   
                            sell_result = buylimit(exchange, market, sell_q_step, price_to_sell, balance_available) 
                        # print "selllimit({}, {}, {}, {}, {})".format(exchange, market, sell_q_step, price_to_sell, balance_available) #DEBUG 
                    else: 
                        sell_result = selllimit(exchange, market, sell_q_step, price_to_sell) 
                    
                    logger.lprint(["-------------------------------------------------------------------- \n>> Sell result:", sell_result, "\n--------------------------------------------------------------------"])   
                    
                    if (sell_result == err_1) or (sell_result == err_2):
                        sell_run_flag = False
                        stopmessage = 'err_low'
                    else:
                        # Checking if the sell order was placed
                        try: 
                            if 'uuid' not in sell_result.keys():
                                # Issue with placing order
                                # DEBUG # print "Issue"
                                sell_run_flag = False
                                stopmessage = 'no_idea'
                        except:
                            # DEBUG # print "Issue"
                            sell_run_flag = False
                            stopmessage = 'no_idea'
                
                else: 
                    # If in simulation - calculate profit from virtual sale.  
                    if exchange != 'bitmex': 
                        main_curr_from_sell += float(sell_q_step) * price_to_sell 
                        commission_total += float(sell_q_step)*price_to_sell * comission_rate
                    else: 
                        main_curr_from_sell += contracts_start/price_to_sell
                        commission_total = 0 
                    sell_run_flag = False  
                    
                # Update the db with price_last_sell
                sql_string = "UPDATE jobs SET price_curr={}, selling={} WHERE job_id={}".format(round(price_last_sell, 8), 1, job_id)
                rows = query(sql_string)

                # Decrease remaining balance to sell 
                remaining_sell_balance = remaining_sell_balance - sell_q_step

        # Checking Telegram requests and answering 
        approved_flag = check_cancel_flag()
        if approved_flag == False: 
            # Aborting if asked
            sell_run_flag = False
            stopmessage = 'abort_telegram'
        # Change the flag to sleep on the next cycle
        proceed_w_sleep = True    
        
    # Finishing up
    #print "main_curr_from_sell {}, commission_total {}, contracts_start {}".format (main_curr_from_sell,  commission_total, contracts_start) # DEBUG 
    
    # For bitmex 
    bitmex_sell_avg_arr_np = np.array(bitmex_sell_avg_arr)
    bitmex_sell_avg = bitmex_sell_avg_arr_np.mean()
    
    return stopmessage

################################ Functions - system ############################################
def check_cancel_flag():
    global db, cur, job_id
    keep_running = True 
    sql_string = "SELECT abort_flag FROM buys WHERE job_id = '{}'".format(job_id)
    rows = query(sql_string)
    try: 
        flag_terminate = rows[0][0] # first result 
    except: 
        flag_terminate = 0
    if (flag_terminate == 1): 
        keep_running = False
    return keep_running

# Checking if we need to initiate selling from the main or from the mooning cycle 
def check_sell_flag():
    global market 
    global db, cur, job_id
    
    sell_initiate = False 
    sql_string = "SELECT selling FROM jobs WHERE market = '{}'".format(market)
    rows = query(sql_string)

    try: 
        sell_flag = rows[0][0] # first result 
    except: 
        sell_flag = 0
    if (sell_flag == 1): 
        sell_initiate = True
    return sell_initiate
       
def timenow():
    return strftime("%Y-%m-%d %H:%M:%S", localtime())

###################################################################################
###################################################################################    
###################################################################################
############################## MAIN WORKFLOW #####################################
###################################################################################
###################################################################################
###################################################################################

### Greetings (for logs readability) 
logger.lprint(["###################### ROBOT ###########################"])

if stop_loss: 
    logger.lprint([market, "| Take profit target:", price_target, "| Stop loss:", sl_target, "| Simulation mode:", simulation])
else: 
    logger.lprint([market, "| Take profit target:", price_target, "| Stop loss: disabled (post-profit only) | Simulation mode:", simulation])

if limit_sell_amount > 0: 
    logger.lprint(["Maximum quantity to sell", limit_sell_amount])

### Set up the margin on bitmex 
if exchange == 'bitmex': 
    try: # can give an error  
        set_margin = bitmex_leverage(market, bitmex_margin)
    except: 
        logger.lprint(["Cannot set the margin, check manually"])
        chat.send("Cannot set the margin on bitmex, check manually")

'''
######## Removed this - will not really be using SL (at least for now), also not applicable for shorts 
# Check if TP is set higher than SL 
if tp < sl: 
    # print "TP {}, SL {}".format(tp, sl) # DEBUG #
    logger.lprint(["Take profit lower than stop loss, r u ok?"])
    exit(0)
''' 
 
time_hour = time.strftime("%H")     # For periodic updates of TD candles and stops 
time_hour_comms = time_hour     # For periodic status updates 
  
# 1. Checking market correctness and URL validity, as well as protecting from fat fingers
try: 
    ticker_upd = coinigy.price(exchange_abbr, market)
    # Ticker could be failing if there is automatic maintenance - then sleep for a while
    if ticker_upd is None: 
        aux_functions.send_notification(market, chat, 'Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        while ticker_upd is None: 
            logger.lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            time.sleep(300) # sleeping for 5 minutes and checking again
            ticker_upd = coinigy.price(exchange_abbr, market)
        
    if ticker_upd == 'INVALID_MARKET': 
        logger.lprint(['Error: Invalid market'])
        logger.close_and_exit()

    else:
        # Fat fingers protection    
        price_check = ticker_upd
        ratio = float(price_target)/float(price_check)
        if (ratio >= 8) or (ratio <= 0.15): 
            logger.lprint(['Error: Double-check prices, are you missing a zero or adding an extra one? The current price is', price_check])
            logger.close_and_exit()

except urllib2.URLError:
    aux_functions.terminate_w_message(logger, 'Exchange unavailable', 'Exchange url unavailable')
    logger.close_and_exit()


### 2. Checking available balance. If bitmex - checking whether we have a long or a short position 
if simulation != True: 
    balance = getbalance(exchange, currency)
    if balance['Available'] == 0: 
        aux_functions.terminate_w_message(logger, 'Error: Zero balance', currency + ': zero balance')
        logger.close_and_exit()
    if exchange == 'bitmex': 
        bitmex_no_positions = True 
        positions = getpositions(exchange, market) 
        for position in positions: 
            if position != {}: 
                bitmex_no_positions = False 
                if position['type'] == 'short':     # enabling short flag if the positions are in short 
                    short_flag = True 
                    logger.lprint(['Enabling short_flag'])
                break # exit the for loop 
        if bitmex_no_positions: 
            aux_functions.terminate_w_message(logger, 'Error: Zero positions on bitmex', 'Error: Zero positions on bitmex')
            logger.close_and_exit()
else: # if simulation 
    if price_target < sl_target: 
        short_flag = True
    else: 
        short_flag = False 
            
### 3. Start the main workflow
run_flag = True 
approved_flag = True
no_sell_orders = False      # default value to suit both simulation and the real run

### 4. Inserting in the sqlite db if started fine ##
sql_string = "INSERT INTO jobs(market, tp, sl, simulation, mooning, selling, price_curr, percent_of, abort_flag, stop_loss, entry_price, mode, tp_p, sl_p, exchange) VALUES ('{}', {}, {}, {}, {},  {},  {},  {},  {}, {}, {}, '{}', {}, {}, '{}')".format(
    market.upper(), price_target, sl_target, int(simulation), int(False), int(False), price_curr, 100, int(False), int(stop_loss), price_entry, simulation_param, tp_p, sl_p, exchange)    
job_id, rows = query_lastrow_id(sql_string)

### 5. Price data for time analysis and strategy. Check what's up with TD analysis data 
start_time = time.time()
td_data_available = True  # default which will be changed to False when needed  

try: 
    td_info = tdlib.tdlib()
    bars = td_info.stats(market, exchange_abbr, td_period, 35000, 15, short_flag, market_ref, exchange_abbr_ref)            
    # 35000 entries is enough for 4H and 9H analysis    
    try: 
        if bars == None: 
            td_data_available = False 
    except: 
        for elem in bars['td_setup'][-3:]:      # should have at least 3 bars with filled TD values
            if elem is None: 
                td_data_available = False 
        num_null = bars['open'].isnull().sum()
        if num_null > 0: 
            td_data_available = False 
    del td_info 
except: 
    td_data_available = False 
    
print "TD data availability:", td_data_available
# Changing the flip flag to True if td data is not available 
if not td_data_available: 
    price_flip = True
    
### 7. 4H-based stop loss update    
if td_data_available: 
    logger.lprint(["Reconfiguring stop loss level based on TD candles"])
    price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd = stop_reconfigure('now')
    
    #print ">>> Returned price_flip {}, sl_target_upd {}, sl_upd {}, sl_p_upd {}, sl_extreme_upd {}".format(price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd)  #DEBUG
    if sl_target_upd is not None: 
        sl_target = sl_target_upd
        sl = sl_upd
        sl_p = sl_p_upd    
    if sl_extreme_upd is not None: 
        sl_extreme = sl_extreme_upd
    if price_flip_upd is not None: 
        price_flip = price_flip_upd
        
### 8. Creating new set to store previously executed orders. Will be used to calculate the gains 
orders_start = set()
orders_new = set()

orders_opening = None   #sometimes api fails and ends with an error - so retrying here
while orders_opening is None:
    try:
        orders_opening = getorderhistory(exchange, market)
    except:
         time.sleep(1) 
 
logger.lprint(["Last orders when starting the script"])
if len(orders_opening) < 5: 
    count_max = len(orders_opening)
else: 
    count_max = 5 

for i in range (0, count_max): 
    logger.lprint(['>', orders_opening[i]['OrderUuid']])  
 
for elem in orders_opening: 
    orders_start.add(elem['OrderUuid'])
    #logger.lprint(['>', elem['OrderUuid']]) #DEBUG

# Flags to notify if the prices dropped 
flag_notify_m = True
flag_notify_h = True
dropped_flag = False

# Defining the take profit level on initial entry if enabled
take_profit_reached = False 
if config.take_profit:
    if not short_flag:
        take_profit_price = price_entry * (1 + config.take_profit_threshold)
    else:
        take_profit_price = price_entry * (1 - config.take_profit_threshold)
    logger.lprint(["Take profit price:", take_profit_price])
    take_profit_price_previous = take_profit_price
    time_bars_15min_initiated = time.time()
    # First 15-min bars check run 
    td_info = tdlib.tdlib()
    bars_15min = td_info.stats(market, exchange_abbr, '15min', 1000, 5, short_flag, market_ref, exchange_abbr_ref)         
    del td_info     
    
### 9. Start the main cycle of the robot 
while run_flag and approved_flag:  
    try:    # try & except is here to raise keyboard cancellation exceptions
        if td_data_available:         # update the stop loss level if due and if we have data
            price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd = stop_reconfigure()
            #print ">>> Returned price_flip {}, sl_target_upd {}, sl_upd {}, sl_p_upd {}, sl_extreme_upd {}".format(price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd) #DEBUG 
            if sl_target_upd is not None: 
                sl_target = sl_target_upd
                sl = sl_upd
                sl_p = sl_p_upd    
                sql_string = "UPDATE jobs SET sl={}, sl_p={} WHERE job_id={}".format(sl_target, sl_p, job_id)   # updating the DB 
                rows = query(sql_string)
            if sl_extreme_upd is not None: 
                sl_extreme = sl_extreme_upd
            if price_flip_upd is not None: 
                price_flip = price_flip_upd
        
        # Get the last price
        price_last = coinigy.get_avg_price(exchange, exchange_abbr, market, logger)
        
        price_compared = round((float(price_last)/float(price_curr))*100, 2)
        logger.lprint([exchange, market, ": updating price information:", price_last, "|", price_compared, "% of entry price | sl:", sl_target, "sl_extreme", sl_extreme, "| price flip:", price_flip])
        sql_string = "UPDATE jobs SET price_curr={}, percent_of={} WHERE job_id={}".format(round(price_last, 8), price_compared, job_id)
        rows = query(sql_string)
        
        # Commenting for now because the current approach is always time & price based and the approach should not change when the TP is reached 
        # Uncomment if you do not plan to use price analysis or if you would like to create your custom strategy 
        ''' 
        ### Running the main conditions check to trigger take profit / stop loss 
        ## Checking TP reached - for longs / shorts 
        if (short_flag and (price_last < price_target)) or ((not short_flag) and (price_last > price_target)):      
            logger.lprint(["Take-profit price reached"])
            #send_notification("Mooning", market + ": Reached the initial TP target and mooning now: " + str(price_last))
            status = to_the_moon(price_last)    # mooning for as long as possible 
            if status == 'stop':
                logger.lprint(["Stopped monitoring and finished trades (post-profit)"])
                sleep_timer = 0
                run_flag = False 
                stopped_mode = 'post-profit'    # used in buyback
            elif status == 'abort_telegram': 
                logger.lprint(["Stopped monitoring and finished trades (as requested)"])
                sleep_timer = 0
                run_flag = False 
                stopped_mode = 'telegram' 
        ''' 
                
        ## Pierced through stop loss with flip if stop loss is enabled 
        if stop_loss: 
            # Checking for sequential candle-based signals and exits when there is a flip and an opposite-color candles start trading below each other 
            if (short_flag and price_flip and (price_last >= sl_target)) or ((not short_flag) and price_flip and (price_last <= sl_target)):      
            #if (short_flag and price_flip and (price_last >= sl_target/2)) or ((not short_flag) and price_flip and (price_last <= sl_target*2)):        #TEST 10 MIN 
                dropped_flag = True     # changing the flag 
                logger.lprint(["Hitting stop loss threshold:", sl_target])
                # Check if we need to sell 
                if not td_data_available: 
                    sale_trigger = ensure_sale(sl_target)   
                else: 
                    # We will be checking here if last 10min candles closed above the threshold 
                    sale_trigger = ensure_td_sale(sl_target)                   
                logger.lprint(["Sale (stop) trigger:", sale_trigger])
                if sale_trigger: 
                    chat.send(market +": exiting based on the time candles (shorter period)")
                 
            # Checking for extreme moves 
            if sl_extreme is not None and not sale_trigger:      # not sale trigger because we are checking this after the 4h-based candles price
                if (short_flag and (price_last > sl_extreme)) or (not short_flag and (price_last < sl_extreme)): 
                    # Check if we need to sell
                    if not td_data_available: 
                        sale_trigger = ensure_sale(price_last)   
                    else: 
                        sale_trigger = True             
                    logger.lprint(["Sale trigger", sale_trigger])
                    comm_string = "{}: exiting based on the breach of extreme {}. Current price {}".format(market, sl_extreme, price_last)
                    chat.send(comm_string)
                    logger.lprint([comm_string])
        
            # Checking for nines and breaches of nines
            # Approach A: if [-2] (before last) is nine and the price goes beyond the extreme of nine + (or -) contingency
            # Approach B: to take profit just right after the TD 9, and then re-evaluate (currently implemented)     
            # bars_4h is a global var in the stop_reconfigure procedure
            
            ''' 
            # Approach A
            if td_data_available and not sale_trigger:     
                if bars_4h['td_setup'].iloc[-2] == '9': 
                    if short_flag: #shorts 
                        if price_last > bars_4h['high'].iloc[-2]*(1 + var_contingency):
                            sale_trigger = True   
                    else: #longs 
                        if price_last < bars_4h['low'].iloc[-2]*(1 - var_contingency):
                            sale_trigger = True      
                if sale_trigger: 
                    comm_string = "{}: exiting based on 9-1 rule. Current price {}".format(market, price_last)
                    logger.lprint([ comm_string ])
                    chat.send(comm_string)
            '''     
            # Approach B 
            if td_data_available and not sale_trigger:      
                if bars_4h['td_setup'].iloc[-2] == '9': 
                    sale_trigger = True    
                    comm_string = "{}: exiting after 9 (taking profit for re-evaluation further). Current price {}".format(market, price_last)
                    logger.lprint([ comm_string ])
                    chat.send(comm_string)
            
            # Checking for RSI conditions 
            # If short: take profit on 4h RSI < 26 and 1h RSI < 30, then usual reentry; 
            # If long: 4h rsi >= 83.5, and 1h >= 81.5 then usual reentry; 
            # these rules are empirical, find better values if you could 
            # Variables rsi_4h, rsi_1h
            
            if td_data_available and not sale_trigger:      
                if ( ((short_flag) and (rsi_4h <= 26.5) and (rsi_1h <= 30)) or  
                    ((not short_flag) and (rsi_4h >= 83.5) and (rsi_1h >= 81.5)) ): 
                    sale_trigger = True             
                    comm_string = "{}: taking profit based on extreme RSI value".format(market)
                    chat.send(comm_string)
                    logger.lprint([comm_string])
            
            # Checking whether a take_profit_price should be updated. Using price data on 15-min candles 
            if td_data_available and config.take_profit: 
                td_info = tdlib.tdlib()
                time_bars_15min_elapsed = (math.ceil(time.time() - time_bars_15min_initiated ))/60    
                if time_bars_15min_elapsed >= 15: 
                    bars_15min = td_info.stats(market, exchange_abbr, '15min', 1000, 5, short_flag, market_ref, exchange_abbr_ref)     
                    time_bars_15min_initiated = time.time()
                
                del td_info 
                
                if (not short_flag) and (price_last > take_profit_price):
                    # If long and higher, reconfiguring the stop to be 15-min candle open (before the current) 
                    take_profit_price_previous = take_profit_price
                    take_profit_price = bars_15min['low'].iloc[-2]
            
                if (short_flag) and (price_last < take_profit_price):
                    take_profit_price_previous = take_profit_price
                    take_profit_price = bars_15min['high'].iloc[-2]
                
                if take_profit_price != take_profit_price_previous: 
                    logger.lprint(["> Take profit price update:", take_profit_price])
                    take_profit_reached = True 
  
            # Checking whether we should take profit now. Only executing after the initial take profit target is reached 
            # Accounting for 0.25% margin of error / noise 
            if config.take_profit and take_profit_reached and not sale_trigger:   
                if ((not short_flag) and (price_last < take_profit_price*0.9975)) or ((short_flag) and (price_last > take_profit_price*1.0025)): 
                    # Time to take some profit 
                    sale_trigger = True             
                    comm_string = "{}: taking profit as per thresholds".format(market)
                    chat.send(comm_string)
                    logger.lprint([comm_string])   
            
            ### Stop loss triggered 
            if sale_trigger == True:       
                # Stop-loss triggered
                logger.lprint(["Triggering stop on", price_last])
                aux_functions.send_notification(market, chat, 'Sell: SL', exchange + ' : ' + market + ': Triggering stop at the level of ' + str(price_last))
                status = sell_now(price_last)
                # Handling results
                run_flag, stat_msg = process_stat(status)
                logger.lprint([stat_msg])
                stopped_mode = 'pre-profit'     # used in buyback
                stopped_price = sl_target          
           
        # Check if selling now request has been initiated
        sell_init_flag = check_sell_flag()
        if sell_init_flag and approved_flag and run_flag:       
            logger.lprint(["Sale initiated via Telegram @", price_last])
            status = sell_now(price_last)
            # Handling results
            run_flag, stat_msg = process_stat(status)
            logger.lprint([stat_msg])
            stopped_price = price_last  # used in buyback
            stopped_mode = 'telegram' 
        
        # Checking cancellation request and sleeping 
        if run_flag and approved_flag:
            approved_flag = check_cancel_flag()
            if approved_flag == False: 
                logger.lprint(["Shutdown was requested via Telegram"])   
                stopped_mode = 'telegram' 
                sleep_timer = 0
            time.sleep(sleep_timer)
            
    except KeyboardInterrupt:
        logger.lprint(["Shutdown was initiated manually, canceling orders and terminating now"])   
        sql_string = "DELETE FROM jobs WHERE job_id = {}".format(job_id)    # deleting the task from the db 
        rows = query(sql_string)

        # Cancelling orders if not in the simulation mode
        if simulation != True:
            cancel_orders(exchange, market)
            time.sleep(10) # wait for cancellations to be processed
            # Information on orders performed if not in a simulation mode
            sell_orders_info()
            sell_orders_outcome()
        logger.close_and_exit()


### 10. Exit point for the main cycle, sell cycle, mooning cycle 
sql_string = "DELETE FROM jobs WHERE job_id = {}".format(job_id)    # deleting the task from the db 
rows = query(sql_string)

# Just a simulation and cancelled by Telegram thing - no virtual sell orders
if simulation == True and approved_flag != True:   
    no_sell_orders = True

### 11. Getting information on performed sell orders and displaying / recording the outcomes
sell_orders_info()
sell_orders_outcome()

# Then start monitoring for buyback (both post-moon and SL)  'pre-profit'  /  'post-profit'
''' 
# Uncomment if you would like to restrict buybacks, code should also be revised from long / short perspective 
# Checking losses to stop buyback in case of 2 consecutive losses incurred  
sql_string = "SELECT id FROM losses WHERE market = '{}'".format(market)
rows = query(sql_string)
try: 
    loss_id = int(rows[0][0]) # if one loss already have been incurred - no bback 
except: 
    loss_id = None
''' 
if (stopped_mode == 'pre-profit') and (cancel_buyback == False): 
    if short_flag != True: # LONGS  
        bb_price = price_exit * 0.9975  # Using value of price_exit (actual sell price) minus 0.25% (for non-TD-based price action only) 
    else: 
        bb_price = price_exit * 1.0025 # SHORTS  
    if td_data_available:  
        logger.lprint(["Buyback will be based on TD candles"])
    else: 
        logger.lprint(["Setting buyback price as actual sell +/- 0.25%. Price_exit:", price_exit, "bb_price", bb_price])
    
    '''
    # Uncomment if you would like to restrict buybacks, code should also be revised from long / short perspective 
    # Checking losses to stop buyback in case of 2 consecutive losses incurred  
    if loss_id is not None: 
        chat.send(market +": stopped buyback after two consecutive losses")
        sql_string = "DELETE FROM losses WHERE id = {}".format(loss_id)
        rows = query(sql_string)
        logger.close_and_exit()
        exit(0) 
    else: 
        # Inserting into losses table if this is the first occurence 
        sql_string = "INSERT INTO losses(market) VALUES ('{}')".format(market)
        rows = query(sql_string)    
    ''' 
elif stopped_mode == 'post-profit': 
    # Thresholds for post-profit fallback for BTC or ALTS
    if market == 'USDT-BTC': 
        if short_flag != True: # LONGS  
            bb_price = price_exit  * 1.005 # Using fixed value of +0.5% from stop; however, does not refer to price value when using TD analysis 
        else: 
            bb_price = price_exit  * 0.995
    else: 
        if short_flag != True: # LONGS  
            bb_price = price_exit  * 1.01  
        else: 
            bb_price = price_exit  * 0.99
     
    if td_data_available:  
        logger.lprint(["Buyback will be based on TD candles"])
    else: 
        logger.lprint(["Setting buyback price as actual +/- 1%. Stopped_price:", stopped_price, "bb_price", bb_price])
    ''' 
    # Uncomment if you would like to restrict buybacks 
    # If it was a loser - delete the info in DB and continue with BBack 
    if loss_id is not None: 
        sql_string = "DELETE FROM losses WHERE id = {}".format(loss_id)
        rows = query(sql_string)
    ''' 
else: 
    # If just called for stop from Telegram 
    logger.lprint(["Sold through telegram"])   
    if price_exit is None: 
        bb_price = price_last
    else: 
        bb_price = price_exit

### 12. Buying back based on 4H action or alternative price action    
try: 

    # Starting buyback except for the cases when the task was aborted through telegram #Commented now - reconsidered that I still need this  
    #if stopped_mode != 'telegram':      
    logger.lprint(["Buyback monitoring started:", stopped_mode, "| TD data availability:", td_data_available, "| Ride pullbacks:", config.ride_pullbacks ])   
    
    if exchange == 'bitmex': 
        buy_trade_price = (main_curr_from_sell)/bitmex_margin
    else: 
        buy_trade_price = float(balance_start) * bb_price * (1 - comission_rate)    # commission depending on the exchange. If we do not have TD data
    
    # Inserting into buyback information table 
    sql_string = "INSERT INTO bback(market, bb_price, curr_price, trade_price, exchange) VALUES ('{}', {}, {}, {}, '{}')".format(market, bb_price, bb_price, buy_trade_price, exchange)
    bb_id, rows = query_lastrow_id(sql_string)      
    
    time_bb_initiated = time.time()     # getting a snapshot of time for buyback so that we wait for at least an hour before starting buyback 
    bb_flag, direction = buy_back(bb_price)      # runs until a result is returned with a confirmation and a direction which defines the next step  
    
    # If we have reached the target to initiate a buyback and there was no cancellation through Telegram
    if bb_flag: 
        aux_functions.send_notification(market, chat, 'Buyback', 'Buy back initiated for ' + market + ' on ' + exchange + '. Direction: ' + direction)  
        
        # Launching workflow to buy and resume the task with same parameters
        # Insert a record in the db: workflow(wf_id INTEGER PRIMARY KEY, tp FLOAT, sl FLOAT, sell_portion FLOAT)
        
        if direction == 'up': #LONGS 
            sl_price = bb_price * (1 - var_contingency)  # depending on the strategy 
            tp_price = bb_price * tp
        elif direction == 'down': 
            sl_price = bb_price * (1 + var_contingency)  # depending on the strategy 
            tp_price = bb_price / tp

        #print "bb_price {}, tp {}, var_contingency {}, tp_price {}, sl_price {}".format(bb_price, tp, var_contingency, tp_price, sl_price) # DEBUG 
            
        sql_string = "INSERT INTO workflow(tp, sl, sell_portion, run_mode, price_entry, exchange) VALUES ({}, {}, {}, '{}', {}, '{}')".format(tp_price, sl_price, 0, simulation_param, float(bb_price), exchange_abbr)
        wf_id, rows = query_lastrow_id(sql_string)       

        if wf_id is not None: 
            buy_market = '{0}-{1}'.format(trade, currency)
            sql_string = "UPDATE workflow SET market = '{}', trade = '{}', currency = '{}', exchange = '{}' WHERE wf_id = {}".format(market, trade, currency, exchange_abbr, wf_id) 
            job_id, rows = query_lastrow_id(sql_string)
            
        # Buy depending on the platform. We will buy @ market price now, and the price entry price is already in the DB
        if td_data_available: 
            mode_buy = 'now' 
        else: 
            mode_buy = 'reg' 

        logger.close() # closing logs 
        
        sql_string = "DELETE FROM bback WHERE id = {}".format(bb_id)  # deleting buyback from the table 
        rows = query(sql_string)
        
        # Run a smart buy task now when we have a buyback confirmation 
        if direction == 'up': #LONGS 
            python_call = 'python smart_buy.py ' + ' '.join([mode_buy, exchange_abbr, trade + '-' + currency, str(buy_trade_price)])
        elif direction == 'down': #SHORTS - need to change plus to minus
            python_call = 'python smart_buy.py ' + ' '.join([mode_buy, exchange_abbr, trade + '-' + currency, str(-buy_trade_price)])
        print '>>>' + python_call
        # Open in the same window 
        p = subprocess.Popen(python_call, shell=True, stderr=subprocess.PIPE)
        while True:
            out = p.stderr.read(1)
            if out == '' and p.poll() != None:
                break
            if out != '':
                sys.stdout.write(out)
                sys.stdout.flush()
    
    # If a buyback cancellation was requested 
    else: 
        aux_functions.send_notification(market, chat, 'Buyback', 'Buy back cancelled as requested for ' + market + ' on ' + exchange) 
        # Delete buyback from the DB 
        sql_string = "DELETE FROM bback WHERE id = {}".format(bb_id)
        rows = query(sql_string)
    
    # If telegram stop - just finalise and exit  - commented 
    #else:  
    #    logger.close_and_exit()
    
except KeyboardInterrupt:
    print "Buyback cancelled or the task was finished"  
     # Delete buyback from the DB 
    sql_string = "DELETE FROM bback WHERE id = {}".format(bb_id)
    rows = query(sql_string)
    try: 
        logger.close_and_exit()
    except: 
        print 'Logs already closed' 
